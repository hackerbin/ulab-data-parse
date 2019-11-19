import json
import os
import sys

import openpyxl
import xlsxwriter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
args = sys.argv
file_name = 'response.xlsx'
if len(args) > 1:
    file_name = args[1]

xcell_object = openpyxl.load_workbook(os.getcwd() + '/' + file_name)  # edit file name and path as you need
worksheet = xcell_object["Form Responses 1"]  # edit sheet name as you need
print(worksheet)
excel_data = list()
json_data = []
for row in worksheet.iter_rows():
    row_data = list()
    for cell in row:
        row_data.append(str(cell.value))
    excel_data.append(row_data)
    exdata = {
        # 'time': row_data[0] if row_data[0] != None or row_data[0] != 'None' else None,
        # 'email': row_data[1] if row_data[1] != None or row_data[1] != 'None' else None,
        'name': row_data[2] if row_data[2] != None or row_data[2] != 'None' else None,
        # 'address': row_data[3] if row_data[3] != None or row_data[3] != 'None' else None,
        # 'phone': row_data[4] if row_data[4] != None or row_data[4] != 'None' else None,
        # 'gender': row_data[5] if row_data[5] != None or row_data[5] != 'None' else None,
        'batch': str(row_data[6])[:3] if row_data[6] != None or row_data[6] != 'None' else None,
        'current_position': row_data[7] if row_data[7] != None or row_data[7] != 'None' else None,
        'msc_bd': row_data[8] if row_data[8] != None or row_data[8] != 'None' else None,
        'msc_abroad': row_data[9] if row_data[9] != None or row_data[9] != 'None' else None,
        'company_and_post': row_data[10] if row_data[10] != None or row_data[10] != 'None' else None,
        'own_company_and_address': row_data[11] if row_data[11] != None or row_data[11] != 'None' else None,
        'other': row_data[12] if row_data[12] != None or row_data[12] != 'None' else None,
        'company': row_data[10].split(',')[0],
        'position': row_data[10].split(',')[1] if len(row_data[10].split(',')) > 1 else None
    }
    company = row_data[10]
    # print(exdata)
    json_data.append(exdata)
    # json_data.pop()
del json_data[0]
with open('data.json', 'w') as outfile:
    json.dump(json_data, outfile)

    # write xcel file with company name and post and name
#
workbook = xlsxwriter.Workbook('companies.xlsx')
worksheet = workbook.add_worksheet()
bold = workbook.add_format({'bold': True})
worksheet.set_column('A:A', 40)
worksheet.set_column('B:B', 40)
worksheet.set_column('C:C', 40)
worksheet.set_column('C:C', 15)
worksheet.write('A1', 'Company', bold)
worksheet.write('B1', 'Post', bold)
worksheet.write('C1', 'Name', bold)
worksheet.write('D1', 'Batch', bold)
pos = 2
for i, item in enumerate(json_data):
    if item['company'] and not item['company'] == 'None':
        worksheet.write('A' + str(pos), item['company'])
        worksheet.write('B' + str(pos), item['position'])
        worksheet.write('C' + str(pos), item['name'])
        worksheet.write('D' + str(pos), item['batch'])
        pos += 1
        print(i, item)

workbook.close()
